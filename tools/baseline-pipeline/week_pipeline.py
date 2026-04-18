from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_BACKEND_DIR = Path(__file__).resolve().parents[2] / "backend"
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

from pdf_extract.pipeline import parse_pdf_to_records
from pipeline_config import ACTIVE_STORES, EXCLUDED_STORES, ONE_SIZE_LABEL, SIZE_ORDER, STORE_SALES_RANK, STORE_SIZES
from week_sources import load_store_totals_with_source, resolve_pdf_dir


def normalize_excel_size(header: Any) -> str | None:
    if header is None:
        return None
    text = str(header).strip()
    if not text:
        return None
    return text.split("/")[0].upper()


def load_moves_from_excel(excel_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    size_columns: dict[int, str] = {}
    for idx, header in enumerate(headers, start=1):
        if idx <= 3:
            continue
        size = normalize_excel_size(header)
        if size:
            size_columns[idx] = size

    moves: list[dict[str, Any]] = []
    article_ids: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        article_id = str(row[0]).strip()
        from_store = str(row[1]).strip()
        to_store = str(row[2]).strip()
        article_ids.append(article_id)

        for col_idx, size in size_columns.items():
            raw_qty = row[col_idx - 1]
            if raw_qty in (None, "", 0):
                continue
            try:
                qty = int(raw_qty)
            except (TypeError, ValueError):
                print(
                    f"[WAARSCHUWING] Niet-numerieke hoeveelheid overgeslagen: "
                    f"artikel {article_id}, maat {size}, waarde={raw_qty!r}"
                )
                continue
            if qty <= 0:
                continue
            moves.append(
                {
                    "article_id": article_id,
                    "from_store": from_store,
                    "to_store": to_store,
                    "size": size,
                    "qty": qty,
                }
            )

    unique_article_ids = sorted(set(article_ids), key=lambda value: int(value))
    return moves, unique_article_ids


def find_pdf_files(pdf_dir: Path, article_ids: list[str]) -> tuple[dict[str, Path], list[str], list[str]]:
    pdf_by_article: dict[str, Path] = {}
    for pdf_path in pdf_dir.rglob("*.pdf"):
        pdf_by_article[pdf_path.stem] = pdf_path

    missing = [article_id for article_id in article_ids if article_id not in pdf_by_article]
    extra = sorted(
        [article_id for article_id in pdf_by_article if article_id not in set(article_ids)],
        key=int,
    )
    matched = {article_id: pdf_by_article[article_id] for article_id in article_ids if article_id in pdf_by_article}
    return matched, missing, extra


def sort_sizes(sizes: list[str]) -> list[str]:
    upper_sizes = [size.upper() for size in sizes]
    if all(size.isdigit() for size in upper_sizes):
        return sorted(upper_sizes, key=int)

    indexed = {size: idx for idx, size in enumerate(SIZE_ORDER)}

    def sort_key(size: str) -> tuple[int, str]:
        if "/" in size:
            left = size.split("/", 1)[0]
            return (indexed.get(left, 999), size)
        return (indexed.get(size, 999), size)

    return sorted(upper_sizes, key=sort_key)


def calc_series_width(stock_by_size: dict[str, int], size_order: list[str]) -> int:
    best = 0
    current = 0
    for size in size_order:
        if stock_by_size.get(size, 0) > 0:
            current += 1
            best = max(best, current)
        else:
            current = 0
    return best


def size_position(size: str, size_order: list[str]) -> str:
    if size not in size_order:
        return "unknown"
    idx = size_order.index(size)
    if idx == 0 or idx == len(size_order) - 1:
        return "edge"
    return "middle"


def store_features(stock_by_size: dict[str, int], sold: int, size_order: list[str]) -> dict[str, Any]:
    total_stock = sum(stock_by_size.values())
    missing_sizes = [size for size in size_order if stock_by_size.get(size, 0) == 0]
    double_sizes = sum(1 for qty in stock_by_size.values() if qty >= 2)
    return {
        "total_stock": total_stock,
        "sold": sold,
        "series_width": calc_series_width(stock_by_size, size_order),
        "missing_sizes": missing_sizes,
        "double_size_count": double_sizes,
        "has_min_3_items": total_stock >= 3,
        "has_min_3_sizes": sum(1 for qty in stock_by_size.values() if qty > 0) >= 3,
    }


def normalize_moves_for_snapshot(article_moves: list[dict[str, Any]], snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    if snapshot.get("sizes") == [ONE_SIZE_LABEL]:
        return [
            {
                **move,
                "size": ONE_SIZE_LABEL,
            }
            for move in article_moves
        ]
    return article_moves


def parse_pdf_snapshot(pdf_path: Path, store_totals: dict[str, int]) -> dict[str, Any]:
    parsed = parse_pdf_to_records(str(pdf_path))
    article_id = parsed.meta.get("Volgnummer") or pdf_path.stem
    size_order = sort_sizes(parsed.sizes)

    # DRT table detection can occasionally lock onto the header timestamp.
    # When that happens, rerun the text fallback directly from the raw page text.
    def is_valid_size_token(size: str) -> bool:
        if size == ONE_SIZE_LABEL:
            return True
        if size in SIZE_ORDER:
            return True
        if "/" in size:
            return True
        return size.isdigit() and 30 <= int(size) <= 60

    if parsed.errors or not size_order or not all(is_valid_size_token(size) for size in size_order):
        import pdfplumber
        from pdf_extract.text_parser import parse_from_text_lines

        with pdfplumber.open(str(pdf_path)) as pdf:
            text = pdf.pages[0].extract_text() or ""
        sizes, rows, totals, difference, negative_voorraad = parse_from_text_lines(text)
        if sizes:
            parsed.sizes = sizes
            parsed.rows = rows
            parsed.totals = totals
            parsed.difference = difference
            parsed.negative_voorraad_detected = negative_voorraad
            parsed.errors = []
            size_order = sort_sizes(parsed.sizes)

    stores: dict[str, dict[str, Any]] = {}
    for row in parsed.rows:
        store_code = str(row.get("filiaal_code", "")).strip()
        if not store_code or store_code in EXCLUDED_STORES or store_code not in ACTIVE_STORES:
            continue

        stock_by_size = {
            str(size).upper(): int(qty)
            for size, qty in row.get("voorraad_per_maat", {}).items()
        }
        sold = int(row.get("verkocht", 0) or 0)
        stores[store_code] = {
            **{size: stock_by_size.get(size, 0) for size in size_order},
            "sold": sold,
            "store_name": ACTIVE_STORES[store_code],
            "store_size_class": STORE_SIZES.get(store_code),
            "store_sales_rank": STORE_SALES_RANK.get(store_code),
            "store_total_stock_week": store_totals.get(store_code),
        }

    return {
        "article_id": article_id,
        "pdf_file": str(pdf_path),
        "meta": parsed.meta,
        "sizes": size_order,
        "stores": stores,
        "errors": parsed.errors,
        "warnings": parsed.warnings,
        "negative_voorraad_detected": parsed.negative_voorraad_detected,
    }


def build_combined_record(
    snapshot: dict[str, Any],
    article_moves: list[dict[str, Any]],
    week: int,
    year: int,
    store_totals: dict[str, int],
) -> dict[str, Any]:
    size_order = snapshot["sizes"]
    stores = snapshot["stores"]

    feature_rows: list[dict[str, Any]] = []
    for move in article_moves:
        donor = stores.get(move["from_store"], {})
        receiver = stores.get(move["to_store"], {})
        donor_stock = {size: int(donor.get(size, 0)) for size in size_order}
        receiver_stock = {size: int(receiver.get(size, 0)) for size in size_order}
        donor_store_features = store_features(donor_stock, int(donor.get("sold", 0)), size_order)
        receiver_store_features = store_features(receiver_stock, int(receiver.get("sold", 0)), size_order)

        size = move["size"].upper()
        feature_rows.append(
            {
                "move": move,
                "features": {
                    "donor_stock_for_size": donor_stock.get(size, 0),
                    "receiver_stock_for_size": receiver_stock.get(size, 0),
                    "donor_sales": int(donor.get("sold", 0)),
                    "receiver_sales": int(receiver.get("sold", 0)),
                    "sales_difference": int(receiver.get("sold", 0)) - int(donor.get("sold", 0)),
                    "size_position": size_position(size, size_order),
                    "is_double_size_at_donor": donor_stock.get(size, 0) >= 2,
                    "donor_series_width": donor_store_features["series_width"],
                    "receiver_series_width": receiver_store_features["series_width"],
                    "receiver_missing_size_before": receiver_stock.get(size, 0) == 0,
                    "donor_would_hit_zero": max(donor_stock.get(size, 0) - move["qty"], 0) == 0,
                    "donor_below_min_items_after": max(donor_store_features["total_stock"] - move["qty"], 0) < 3,
                    "receiver_store_rank": receiver.get("store_sales_rank"),
                    "donor_store_rank": donor.get("store_sales_rank"),
                    "donor_store_total_stock_week": donor.get("store_total_stock_week"),
                    "receiver_store_total_stock_week": receiver.get("store_total_stock_week"),
                    "store_total_stock_tiebreak_advantage": (
                        (donor.get("store_total_stock_week") or 0) - (receiver.get("store_total_stock_week") or 0)
                    ),
                },
            }
        )

    return {
        "year": year,
        "week": week,
        "article_id": snapshot["article_id"],
        "snapshot": snapshot,
        "moves": article_moves,
        "move_features": feature_rows,
    }


def ensure_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def run_week_pipeline(
    week: int,
    year: int,
    week_dir: Path,
    output_root: Path,
    excel_path: Path | None = None,
) -> dict[str, Any]:
    excel_path = excel_path or (week_dir / "Ingevoerd.xlsx")
    pdf_dir = resolve_pdf_dir(week_dir)
    if pdf_dir is None:
        raise FileNotFoundError(
            f"Geen PDF-bronmap gevonden in {week_dir}. Verwacht een van: Interfiliaallijsten, print pdf"
        )

    store_totals, store_totals_source = load_store_totals_with_source(week_dir)

    moves, article_ids = load_moves_from_excel(excel_path)
    pdf_files, missing_pdfs, extra_pdfs = find_pdf_files(pdf_dir, article_ids)

    moves_by_article: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for move in moves:
        moves_by_article[move["article_id"]].append(move)

    normalized_moves: list[dict[str, Any]] = []
    snapshots: list[dict[str, Any]] = []
    combined: list[dict[str, Any]] = []
    parse_failures: list[dict[str, Any]] = []

    for article_id in article_ids:
        pdf_path = pdf_files.get(article_id)
        if not pdf_path:
            normalized_moves.extend(moves_by_article.get(article_id, []))
            continue

        snapshot = parse_pdf_snapshot(pdf_path, store_totals=store_totals)
        snapshot["week"] = week
        snapshot["year"] = year
        snapshots.append(snapshot)

        article_moves = normalize_moves_for_snapshot(moves_by_article.get(article_id, []), snapshot)
        normalized_moves.extend(article_moves)

        if snapshot["errors"]:
            parse_failures.append({"article_id": article_id, "errors": snapshot["errors"]})

        combined.append(
            build_combined_record(
                snapshot=snapshot,
                article_moves=article_moves,
                week=week,
                year=year,
                store_totals=store_totals,
            )
        )

    output_dir = output_root / str(year) / f"week_{week}"
    ensure_json(output_dir / "snapshots.json", snapshots)
    ensure_json(output_dir / "moves.json", normalized_moves)
    ensure_json(output_dir / "combined.json", combined)

    analysis = {
        "year": year,
        "week": week,
        "article_count_in_excel": len(article_ids),
        "move_count": len(moves),
        "matched_pdf_count": len(pdf_files),
        "missing_pdf_count": len(missing_pdfs),
        "missing_pdf_article_ids": missing_pdfs,
        "extra_pdf_count": len(extra_pdfs),
        "extra_pdf_article_ids": extra_pdfs,
        "pdf_source_dir": str(pdf_dir),
        "store_totals_loaded_count": len(store_totals),
        "store_totals_source": store_totals_source,
        "store_totals_missing_codes": [
            code for code in ACTIVE_STORES
            if code not in store_totals
        ],
        "parse_failure_count": len(parse_failures),
        "parse_failures": parse_failures,
    }
    ensure_json(output_dir / "analysis.json", analysis)
    return {
        "analysis": analysis,
        "moves": moves,
        "snapshots": snapshots,
        "combined": combined,
        "output_dir": str(output_dir),
    }
