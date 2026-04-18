from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline_config import STORE_NAME_TO_CODE

PDF_DIR_CANDIDATES = ("Interfiliaallijsten", "print pdf")
STORE_TOTAL_FILES = ("store_totals.tsv", "store_totals.txt", "store_totals.json")
PRINT_NUMMERS_WORKBOOK = "Print_nummers.xlsx"
PRINT_NUMMERS_SHEET = "Info + VVN"
PRINT_NUMMERS_MIN_STORES = 6  # Minimaal aantal filialen dat gevonden moet worden


def resolve_pdf_dir(week_dir: Path) -> Path | None:
    for dirname in PDF_DIR_CANDIDATES:
        candidate = week_dir / dirname
        if candidate.exists() and candidate.is_dir():
            return candidate
    return None


def _scan_sheet_for_store_totals(ws: Any) -> dict[str, int]:
    """Scan het werkblad op bekende filiaalcodes en de aangrenzende totaalwaarden.

    Kijkt per gevulde cel of de waarde overeenkomt met een bekende filiaalcode of
    filiaalnaam. Als dat zo is, wordt de direct aangrenzende cel (rechts of links)
    gecontroleerd op een geheel getal als totaalwaarde.
    """
    from pipeline_config import ACTIVE_STORES, STORE_NAME_TO_CODE

    active_codes = set(ACTIVE_STORES.keys())
    result: dict[str, int] = {}

    for row in ws.iter_rows():
        cells = list(row)
        for i, cell in enumerate(cells):
            store_code = _normalize_store_code(cell.value)
            if store_code not in active_codes:
                # Controleer ook op filiaalnaam
                raw = str(cell.value).strip() if cell.value is not None else ""
                store_code = STORE_NAME_TO_CODE.get(raw.lower())
                if store_code is None:
                    continue

            # Zoek totaal in aangrenzende cellen binnen dezelfde rij (rechts en links)
            for offset in (1, -1):
                j = i + offset
                if 0 <= j < len(cells):
                    total = _normalize_total_value(cells[j].value)
                    if total is not None and total > 0:
                        result[store_code] = total
                        break

    return result


def resolve_store_totals_source(week_dir: Path) -> dict[str, Any]:
    for filename in STORE_TOTAL_FILES:
        candidate = week_dir / filename
        if candidate.exists():
            return {
                "kind": "file",
                "path": candidate,
                "label": str(candidate),
            }

    workbook_path = week_dir / PRINT_NUMMERS_WORKBOOK
    if workbook_path.exists():
        loaded = _scan_store_totals_from_workbook(workbook_path)
        label = f"{workbook_path}:{PRINT_NUMMERS_SHEET}"
        return {
            "kind": "workbook",
            "path": workbook_path,
            "label": label,
            "loaded": loaded,
        }

    return {
        "kind": "none",
        "path": None,
        "label": None,
    }


def load_store_totals_with_source(week_dir: Path) -> tuple[dict[str, int], str | None]:
    source = resolve_store_totals_source(week_dir)
    path: Path | None = source["path"]
    if source["kind"] == "file" and path is not None:
        return _load_store_totals_file(path), source["label"]
    if source["kind"] == "workbook" and path is not None:
        loaded = source.get("loaded")
        if isinstance(loaded, dict):
            return loaded, source["label"]
        return _scan_store_totals_from_workbook(path), source["label"]
    return {}, None


def _load_store_totals_file(path: Path) -> dict[str, int]:
    if path.suffix.lower() == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        loaded: dict[str, int] = {}
        for key, value in raw.items():
            normalized = str(key).strip().lower()
            store_code = key if str(key).isdigit() else STORE_NAME_TO_CODE.get(normalized)
            if store_code:
                loaded[str(store_code)] = int(value)
        return loaded

    loaded = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.replace(",", "\t").split("\t") if part.strip()]
        if len(parts) < 2:
            continue
        left, right = parts[0], parts[1]
        if left.lower() in {"store", "winkel", "filiaal"}:
            continue
        store_code = left if left.isdigit() else STORE_NAME_TO_CODE.get(left.lower())
        if store_code:
            loaded[str(store_code)] = int(right)
    return loaded


def _normalize_store_code(raw_code: Any) -> str | None:
    if raw_code in (None, ""):
        return None

    store_code = str(raw_code).strip()
    if store_code.endswith(".0"):
        store_code = store_code[:-2]
    if store_code.isdigit():
        return store_code

    # Formaat "CODE NAAM" zoals "9 Stein" of "11 Brunssum"
    parts = store_code.split(None, 1)
    if parts and parts[0].isdigit():
        return parts[0]

    return STORE_NAME_TO_CODE.get(store_code.lower())


def _normalize_total_value(raw_total: Any) -> int | None:
    if raw_total in (None, ""):
        return None
    try:
        return int(raw_total)
    except (TypeError, ValueError):
        try:
            return int(float(str(raw_total).strip()))
        except (TypeError, ValueError):
            return None


def _scan_store_totals_from_workbook(path: Path) -> dict[str, int]:
    """Laad store_totals uit Print_nummers.xlsx door het werkblad automatisch te scannen.

    De locatie van de store_totals wisselt wekelijks. In plaats van een hardcoded
    celrangschikking te gebruiken, scant deze functie het hele werkblad op bekende
    filiaalcodes en de aangrenzende totaalwaarden (rechts of links).
    Retourneert een lege dict als er minder dan PRINT_NUMMERS_MIN_STORES gevonden worden.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if PRINT_NUMMERS_SHEET not in wb.sheetnames:
            return {}
        ws = wb[PRINT_NUMMERS_SHEET]
        loaded = _scan_sheet_for_store_totals(ws)
        if len(loaded) < PRINT_NUMMERS_MIN_STORES:
            return {}
        return loaded
    finally:
        wb.close()
