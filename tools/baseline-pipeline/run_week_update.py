from __future__ import annotations

import argparse
import sys
from pathlib import Path
from tkinter import Tk, filedialog

from openpyxl import load_workbook

from dataset_refresh import process_single_week, refresh_all_available_weeks
from week_sources import resolve_pdf_dir, resolve_store_totals_source


def prompt_int(label: str, default: int | None = None) -> int:
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{label}{suffix}: ").strip()
        if not raw and default is not None:
            return default
        try:
            return int(raw)
        except ValueError:
            print("Voer een geldig getal in.")


def ask_yes_no(prompt: str, default: bool = True) -> bool:
    suffix = " [Y/n]" if default else " [y/N]"
    raw = input(f"{prompt}{suffix}: ").strip().lower()
    if not raw:
        return default
    return raw in {"y", "yes", "j", "ja"}


def choose_file_dialog(title: str, initialdir: Path, pattern: str) -> Path | None:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askopenfilename(
            title=title,
            initialdir=str(initialdir),
            filetypes=[("Bestanden", pattern), ("Alle bestanden", "*.*")],
            parent=root,
        )
    finally:
        root.destroy()
    return Path(selected) if selected else None


def resolve_excel_path(week_dir: Path, interactive: bool) -> Path:
    default_path = week_dir / "Ingevoerd.xlsx"
    if default_path.exists():
        return default_path

    print(f"`Ingevoerd.xlsx` niet gevonden in {week_dir}.")
    if not interactive:
        raise FileNotFoundError(f"Excel-bestand ontbreekt: {default_path}")

    if ask_yes_no("Wil je het Excel-bestand aanwijzen?", default=True):
        selected = choose_file_dialog(
            title="Kies Ingevoerd.xlsx",
            initialdir=week_dir,
            pattern="*.xlsx",
        )
        if selected:
            return selected

    while True:
        raw = input("Voer het volledige pad naar Ingevoerd.xlsx in: ").strip().strip('"')
        if not raw:
            continue
        path = Path(raw)
        if path.exists():
            return path
        print("Bestand niet gevonden, probeer opnieuw.")


def ensure_store_totals(week_dir: Path, interactive: bool) -> Path | None:
    source = resolve_store_totals_source(week_dir)
    if source["kind"] == "file":
        return source["path"]
    if source["kind"] == "workbook":
        print(f"Store totals worden gelezen uit {source['label']}.")
        return source["path"]

    if not interactive:
        print("Waarschuwing: geen store totals bestand gevonden.")
        return None

    print("Store totals ontbreken voor deze week.")
    print("Plak regels in het formaat `Winkel<TAB>Aantal` of `Winkel,Aantal`.")
    print("Beëindig met een lege regel.")
    lines: list[str] = []
    while True:
        line = input()
        if not line.strip():
            break
        lines.append(line.rstrip())

    if not lines:
        print("Geen store totals ingevoerd. Verwerking gaat door zonder tie-break data.")
        return None

    target = week_dir / "store_totals.tsv"
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Store totals opgeslagen in {target}")
    return target


def excel_article_ids(excel_path: Path) -> list[str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        values.append(str(value).strip())
    return sorted(set(values), key=lambda value: int(value))


def pdf_article_ids(pdf_dir: Path) -> list[str]:
    return sorted({path.stem for path in pdf_dir.rglob("*.pdf")}, key=lambda value: int(value))


def check_excel_pdf_mismatch(excel_path: Path, pdf_dir: Path) -> dict[str, list[str]]:
    excel_ids = set(excel_article_ids(excel_path))
    pdf_ids = set(pdf_article_ids(pdf_dir))
    return {
        "missing_pdfs_for_excel_ids": sorted(excel_ids - pdf_ids, key=int),
        "extra_pdfs_without_excel_id": sorted(pdf_ids - excel_ids, key=int),
    }


def show_mismatch_warning(mismatch: dict[str, list[str]], interactive: bool, pdf_dir: Path) -> None:
    missing = mismatch["missing_pdfs_for_excel_ids"]
    extra = mismatch["extra_pdfs_without_excel_id"]
    if not missing and not extra:
        print("Excel en PDF's matchen op volgnummer.")
        return

    print(f"Mismatch gedetecteerd tussen `Ingevoerd.xlsx` en `{pdf_dir.name}`.")
    print(f"Volgnummers in Excel zonder PDF: {len(missing)}")
    if missing:
        print(", ".join(missing[:20]))
    print(f"PDF's zonder volgnummer in Excel: {len(extra)}")
    if extra:
        print(", ".join(extra[:20]))

    if interactive and not ask_yes_no("Wil je ondanks deze mismatch doorgaan?", default=True):
        raise SystemExit("Verwerking afgebroken door gebruiker.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=None)
    parser.add_argument("--week", type=int, default=None, help="Verwerk alleen deze week.")
    parser.add_argument(
        "--source-root",
        type=Path,
        default=Path(r"Z:\Herverdeellijsten"),
        help="Root van de bronmappen 'Jaar YYYY/Week N'.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("data"),
        help="Root voor gegenereerde datasets.",
    )
    parser.add_argument(
        "--refresh-all",
        action="store_true",
        help="Verwerk alle beschikbare weken in het jaar en refresh aggregate outputs.",
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Schakel prompts uit en fail direct als input ontbreekt.",
    )
    args = parser.parse_args()

    interactive = not args.non_interactive and sys.stdin.isatty()

    year = args.year
    if year is None:
        year = prompt_int("Jaar", default=2026) if interactive else 2026

    if args.refresh_all:
        result = refresh_all_available_weeks(
            year=year,
            source_root=args.source_root,
            output_root=args.output_root,
        )
        print(result)
        return

    week = args.week
    if week is None:
        if not interactive:
            raise SystemExit("Gebruik `--week` of laat interactieve modus aan.")
        week = prompt_int("Weeknummer")

    week_dir = args.source_root / f"Jaar {year}" / f"Week {week}"
    pdf_dir = resolve_pdf_dir(week_dir)

    if pdf_dir is None:
        raise SystemExit(
            f"PDF-map niet gevonden in {week_dir}. Verwacht een van: Interfiliaallijsten, print pdf"
        )

    excel_path = resolve_excel_path(week_dir, interactive=interactive)
    ensure_store_totals(week_dir, interactive=interactive)

    mismatch = check_excel_pdf_mismatch(excel_path, pdf_dir)
    show_mismatch_warning(mismatch, interactive=interactive, pdf_dir=pdf_dir)

    result = process_single_week(
        year=year,
        week=week,
        source_root=args.source_root,
        output_root=args.output_root,
        excel_path=excel_path,
    )
    print(result)


if __name__ == "__main__":
    main()
